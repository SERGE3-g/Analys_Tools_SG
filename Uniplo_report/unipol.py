import customtkinter
from customtkinter import filedialog
import os
import pandas
import openpyxl
import lxml
import gforms
import shutil
from ctypes import windll
from CTkListbox import *
from datetime import date
from CTkMessagebox import CTkMessagebox
import packaging
import packaging.version
import packaging.specifiers
import packaging.requirements

version = "1.0"
lastVersion, logLevel = version, 0
tempType, tempDetails = "", ""

contabReportBaseFile = "DEFAULT EMPTY REPORT"
contabReportSourceFiles = []

def unipolContabClearSourceFiles():
    global contabReportSourceFiles
    contabReportSourceFiles.clear()
    print(contabReportSourceFiles)
    app.unipolContabSourceFilesListLabel.configure(state=customtkinter.NORMAL)
    app.unipolContabSourceFilesListLabel.delete("0.0", "end")
    app.unipolContabSourceFilesListLabel.configure(state=customtkinter.DISABLED)

def unipolContabSourceFilesAdder():
    global contabReportSourceFiles
    temp = selectFilesPath()
    for filePath in temp:
        contabReportSourceFiles.append([os.path.basename(filePath), filePath])
        print(contabReportSourceFiles)
        app.unipolContabSourceFilesListLabel.configure(state=customtkinter.NORMAL)
        app.unipolContabSourceFilesListLabel.insert(customtkinter.END, "\n" + os.path.basename(filePath))
        app.unipolContabSourceFilesListLabel.configure(state=customtkinter.DISABLED)

def selectFilePath():
    filename = filedialog.askopenfilename(initialdir = "/")
    print(filename)
    return filename

def selectFilesPath():
    filename = filedialog.askopenfilenames(initialdir = "/")
    print(filename)
    return filename

def selectUnipolBaseReportPath():
    global contabReportBaseFile
    temp = selectFilePath()
    if temp != "":
        contabReportBaseFile = temp
    updateContabBaseReportLabel()

def updateContabBaseReportLabel():
    global contabReportBaseFile
    app.unipolContabBaseReportPathEntry.configure(state=customtkinter.NORMAL)
    app.unipolContabBaseReportPathEntry.delete(0, customtkinter.END)
    app.unipolContabBaseReportPathEntry.insert(0, contabReportBaseFile)
    app.unipolContabBaseReportPathEntry.configure(state=customtkinter.DISABLED)

def setUnipolContabBaseReportToDefault():
    global contabReportBaseFile
    contabReportBaseFile = "DEFAULT EMPTY REPORT"
    updateContabBaseReportLabel()

def callback(element, page_index, element_index):
    global tempType
    global tempDetails
    if element.name == 'user':
        return str(getUser())
    if element.name == 'version':
        return str(version)
    if element.name == 'type':
        return tempType
    if element.name == 'details':
        return tempDetails

def onlineLogger(level, type, details):
    global tempType
    global tempDetails
    global logLevel

    if level <= logLevel:
        try:
            tempType = type
            tempDetails = details
            url = "https://docs.google.com/forms/d/e/1FAIpQLScAgMLliRFHoxA_wF18mwO2BKInJJR_vmQdg_SZ7w0oheNXZg/viewform?usp=sf_link"
            form = gforms.Form()
            form.load(url)
            print(form.to_str(indent=2))  # a text representation, may be useful for CLI applications
            form.fill(callback)
            form.submit()
        except:
            print("Logging error")
    else:
        print("Event not to log")


def getRemoteConfig():
    global lastVersion
    global logLevel

    try:
        remoteFile = pandas.read_csv('https://docs.google.com/spreadsheets/d/e/2PACX-1vR6N3dKYFaRytzcc_-dpr7wDiGvhzkENUxcOInLdRMAyX2zQe7cHWIKxgVmBR6mE4_CWbGay4u9FMSq/pub?output=csv')
        lastVersion = remoteFile['last version'].values[0]
        logLevel = int(remoteFile['log level'].values[0])
    except:
        onlineLogger(1, "[ERROR]", "Remote config loading error")
        return version, 0

    print(remoteFile)

    return lastVersion, logLevel

def getUser():
    username = ""
    try:
        username = os.environ.get('USERNAME')
    except:
        username = "User"
        onlineLogger(1, "[ERROR]", "Error loading OS username")
    return username

def last_active_row(input_file):
    wb = openpyxl.load_workbook(input_file, read_only=True)
    ws = wb["File"]

    last_row = ws.max_row + 1

    print("Max row contata")

    last_filled_row = last_row

    print("Inizia il loop")

    for rowNum in reversed(range(last_row)):
        print(rowNum)
        if not ws.cell(row=rowNum, column=16).value is None:
            last_filled_row = rowNum
            print("SI")
            break
        print("NO")

    print("Count finita")
    return last_filled_row

def contabReport():

    app.unipolContabProgressBar.grid()

    global contabReportBaseFile
    global contabReportSourceFiles

    try:
        os.remove('.temp.xlsx')
    except:
        print("Il file già non c'è")



    if contabReportBaseFile == "DEFAULT EMPTY REPORT":
        shutil.copyfile("content/defaultUnipolContabBase.xlsx", '.temp.xlsx')
    else:
        shutil.copyfile(contabReportBaseFile, '.temp.xlsx')

    flussoBase = str('.temp.xlsx')
    csvFilePath = contabReportSourceFiles

    df = pandas.DataFrame()

    print(df)

    print(contabReportBaseFile)

    for csvFile in csvFilePath:
        df = pandas.concat([df, pandas.read_csv(csvFile[1], sep=';', header=None, dtype='unicode')], ignore_index=True)
        print("File concat in df")

    print(df)

    print("CSV LETTI!")

    fileLen = last_active_row(flussoBase)

    with pandas.ExcelWriter(flussoBase, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        print("File open")
        df.to_excel(writer, sheet_name="File", index=False, header=False, startrow=fileLen, startcol=1)


    today = date.today()
    print("Today's date:", today)
    today = today.strftime("%d_%m_%Y")

    nomeFileNew = "PROD_INTERFACCIA_SAP_" + today + ".xlsx"

    try:
        os.remove(nomeFileNew)
    except:
        print("Il file già non c'è")

    os.rename(r'.temp.xlsx', nomeFileNew)

    print("Fatto.")
    app.unipolContabProgressBar.grid_remove()
    CTkMessagebox(title="Info", message="Elaborazione terminata!")


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Report Tool")
        self.geometry("600x600")
        customtkinter.set_appearance_mode("Dark")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # create navigation frame
        self.sidebarNavigationFrame = customtkinter.CTkFrame(self, corner_radius=0)
        self.sidebarNavigationFrame.grid(row=0, column=0, sticky="nsew")
        self.sidebarNavigationFrame.grid_rowconfigure(4, weight=1)

        self.sidebarNavigationTitle = customtkinter.CTkLabel(self.sidebarNavigationFrame, text="Report Tool", compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
        self.sidebarNavigationTitle.grid(row=0, column=0, padx=20, pady=20)

        self.goToHomeButton = customtkinter.CTkButton(self.sidebarNavigationFrame, corner_radius=0, height=40, border_spacing=10, text="Home", fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), anchor="w", command=self.goToHomeButton_event)
        self.goToHomeButton.grid(row=1, column=0, sticky="ew")

        self.goToUnipolContabButton = customtkinter.CTkButton(self.sidebarNavigationFrame, corner_radius=0, height=40, border_spacing=10, text="Unipol - Contabilità", fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), anchor="w", command=self.goToUnipolContabButton_event)
        self.goToUnipolContabButton.grid(row=2, column=0, sticky="ew")

        self.goToUnipolCommissionButton = customtkinter.CTkButton(self.sidebarNavigationFrame, corner_radius=0, height=40, border_spacing=10, text="Unipol - Commissioni", fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), anchor="w", command=self.goToUnipolCommissionButton_event)
        self.goToUnipolCommissionButton.grid(row=3, column=0, sticky="ew")


        # Report Tool HOME FRAME
        self.reportToolHomeFrame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.reportToolHomeFrame.grid_columnconfigure(0, weight=1)

        self.reportToolHomeTitle = customtkinter.CTkLabel(self.reportToolHomeFrame, text="Hello, " + getUser() + "!", compound="left", font=customtkinter.CTkFont(size=15, weight="bold"), anchor="w", justify="left")
        self.reportToolHomeTitle.grid(row=0, column=0, padx=20, pady=(20,0), sticky="ew")


        self.configGridFrame = customtkinter.CTkFrame(self.reportToolHomeFrame, corner_radius=0, fg_color="transparent")
        self.configGridFrame.grid_columnconfigure(0, weight=0)
        self.configGridFrame.grid_columnconfigure(1, weight=1)
        self.configGridFrame.grid(row=1, column=0, padx=20, pady=(20,0), sticky="we")

        self.currentVersionLabel = customtkinter.CTkLabel(self.configGridFrame, text="Current Version: ", compound="left", anchor="w", justify="left")
        self.currentVersionLabel.grid(row=0, column=0, padx=0, pady=(10,0), sticky="ew")

        self.currentVersionValue = customtkinter.CTkLabel(self.configGridFrame, text=str(version), compound="left", anchor="w", justify="left")
        self.currentVersionValue.grid(row=0, column=1, padx=(10,0), pady=(10,0), sticky="ew")

        self.lastVersionLabel = customtkinter.CTkLabel(self.configGridFrame, text="Last version available: ", compound="left", anchor="w", justify="left")
        self.lastVersionLabel.grid(row=1, column=0, padx=0, pady=(10,0), sticky="ew")

        self.lastVersionValue = customtkinter.CTkLabel(self.configGridFrame, text=str(lastVersion), compound="left", anchor="w", justify="left")
        self.lastVersionValue.grid(row=1, column=1, padx=(10,0), pady=(10,0), sticky="ew")

        self.loggingLevelLabel = customtkinter.CTkLabel(self.configGridFrame, text="Remote Logging: ", compound="left", anchor="w", justify="left")
        self.loggingLevelLabel.grid(row=2, column=0, padx=0, pady=(10,0), sticky="ew")

        self.loggingLevelValue = customtkinter.CTkLabel(self.configGridFrame, text="Enabled, Level " + str(logLevel), compound="left", anchor="w", justify="left")
        self.loggingLevelValue.grid(row=2, column=1, padx=(10,0), pady=(10,0), sticky="ew")



        # UNIPOL CONTAB FRAME
        self.unipolContabFrame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.unipolContabFrame.grid_columnconfigure(0, weight=1)


        self.unipolContabTitle = customtkinter.CTkLabel(self.unipolContabFrame, text="Unipol - Report Contabilità", compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
        self.unipolContabTitle.grid(row=0, column=0, padx=20, pady=(20,0), sticky="ew")

        self.unipolContabBaseReportLabel = customtkinter.CTkLabel(self.unipolContabFrame, text="Base Report", justify="left", anchor="w")
        self.unipolContabBaseReportLabel.grid(row=1, column=0, padx=20, pady=(20,0), sticky="ew")
        self.unipolContabBaseReportPathEntry = customtkinter.CTkEntry(self.unipolContabFrame,placeholder_text="18")
        self.unipolContabBaseReportPathEntry.insert(0, "DEFAULT EMPTY REPORT")
        self.unipolContabBaseReportPathEntry.configure(state=customtkinter.DISABLED)
        self.unipolContabBaseReportPathEntry.grid(row=2, column=0, padx=20, pady=0, sticky="ew")

        self.unipolContabBaseReportButtonsFrame = customtkinter.CTkFrame(self.unipolContabFrame, corner_radius=0, fg_color="transparent")
        self.unipolContabBaseReportButtonsFrame.grid_columnconfigure(0, weight=1)
        self.unipolContabBaseReportButtonsFrame.grid_columnconfigure(1, weight=1)
        self.unipolContabBaseReportButtonsFrame.grid(row=3, column=0, padx=(0,0), pady=(10,0), sticky="we")

        self.unipolContabBaseReportSelectPath = customtkinter.CTkButton(self.unipolContabBaseReportButtonsFrame, text="SELECT A CUSTOM FILE", command=selectUnipolBaseReportPath)
        self.unipolContabBaseReportSelectPath.grid(row=0, column=0, padx=(20,0), pady=0, sticky="we")

        self.unipolContabBaseReportSetDefault = customtkinter.CTkButton(self.unipolContabBaseReportButtonsFrame, text="SET TO DEFAULT", command=setUnipolContabBaseReportToDefault)
        self.unipolContabBaseReportSetDefault.grid(row=0, column=1, padx=20, pady=0, sticky="we")

        self.unipolContabSourceFilesLabel = customtkinter.CTkLabel(self.unipolContabFrame, text="Source files", justify="left", anchor="w")
        self.unipolContabSourceFilesLabel.grid(row=4, column=0, padx=20, pady=(20,0), sticky="ew")

        self.unipolContabSourceFilesButtonsFrame = customtkinter.CTkFrame(self.unipolContabFrame, corner_radius=0, fg_color="transparent")
        self.unipolContabSourceFilesButtonsFrame.grid_columnconfigure(0, weight=1)
        self.unipolContabSourceFilesButtonsFrame.grid_columnconfigure(1, weight=1)
        self.unipolContabSourceFilesButtonsFrame.grid(row=5, column=0, padx=(0,0), pady=(0,10), sticky="we")

        self.unipolContabSourceFilesAddFile = customtkinter.CTkButton(self.unipolContabSourceFilesButtonsFrame, text="ADD FILEs", command=unipolContabSourceFilesAdder)
        self.unipolContabSourceFilesAddFile.grid(row=0, column=0, padx=(20,0), pady=0, sticky="we")

        self.unipolContabSourceFilesRemoveSelected = customtkinter.CTkButton(self.unipolContabSourceFilesButtonsFrame, text="CLEAR ALL", command=unipolContabClearSourceFiles)
        self.unipolContabSourceFilesRemoveSelected.grid(row=0, column=1, padx=20, pady=0, sticky="we")

        self.unipolContabSourceFilesListLabel = customtkinter.CTkTextbox(self.unipolContabFrame, width=400, corner_radius=10)
        self.unipolContabSourceFilesListLabel.grid(row=6, column=0, padx=20, pady=0, sticky="nsew")

        self.unipolContabSourceFilesListLabel.configure(state=customtkinter.DISABLED)

        self.unipolContabStartElabButton = customtkinter.CTkButton(self.unipolContabFrame, text="Start Report generation", command=contabReport)
        self.unipolContabStartElabButton.grid(row=7, column=0, padx=20, pady=10, sticky="we")

        self.unipolContabProgressBar = customtkinter.CTkProgressBar(self.unipolContabFrame, mode="indeterminate")
        self.unipolContabProgressBar.grid(row=8, column=0, padx=20, pady=10, sticky="we")
        self.unipolContabProgressBar.start()
        self.unipolContabProgressBar.grid_remove()


        # create third frame
        self.third_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")

        # select default frame
        self.select_frame_by_name("home")

    def select_frame_by_name(self, name):
        # set button color for selected button
        self.goToHomeButton.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.goToUnipolContabButton.configure(fg_color=("gray75", "gray25") if name == "frame_2" else "transparent")
        self.goToUnipolCommissionButton.configure(fg_color=("gray75", "gray25") if name == "frame_3" else "transparent")

        # show selected frame
        if name == "home":
            self.reportToolHomeFrame.grid(row=0, column=1, sticky="nsew")
        else:
            self.reportToolHomeFrame.grid_forget()
        if name == "frame_2":
            self.unipolContabFrame.grid(row=0, column=1, sticky="nsew")
        else:
            self.unipolContabFrame.grid_forget()
        if name == "frame_3":
            self.third_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.third_frame.grid_forget()

    def goToHomeButton_event(self):
        self.select_frame_by_name("home")

    def goToUnipolContabButton_event(self):
        self.select_frame_by_name("frame_2")

    def goToUnipolCommissionButton_event(self):
        self.select_frame_by_name("frame_3")

    def change_appearance_mode_event(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)




if __name__ == "__main__":
    getRemoteConfig()
    onlineLogger(3, "[INFO]", "User logged in")
    app = App()
    app.mainloop()
